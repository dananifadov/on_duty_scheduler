#!/usr/bin/env python3
"""
Excel writer for on-duty scheduler
Creates proper Excel files with colorful formatting, sheets, and blocked days visualization
"""
import json
from pathlib import Path
from datetime import date, datetime, timedelta
import calendar
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class ExcelWriter:
    """Creates proper Excel files with multiple sheets and formatting"""
    
    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Define colors matching duty types and schedule logic
        self.colors = {
            # Assignment type colors
            'WD': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),     # Pastel yellow (weekday)
            'Th': PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid'),     # Pastel orange (Thursday)
            'WE': PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid'),     # Pastel green (weekend)
            'B': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid'),      # Pastel blue (backup)
            'HO': PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid'),     # Pastel purple (holiday)
            
            # Special states
            'blocked': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'), # Pastel red for blocked
            'header': PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid'),  # Pastel blue header
            'empty': PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid'),   # Light gray for empty cells
            
            # Day colors matching duty logic
            'monday': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),    # Same as WD (weekday)
            'tuesday': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),   # Same as WD (weekday)
            'wednesday': PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'), # Same as WD (weekday)
            'thursday': PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid'),  # Same as Th (Thursday)
            'friday': PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid'),    # Same as WE (weekend)
            'saturday': PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid'),  # Same as WE (weekend)
            'sunday': PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid'),    # Same as WE (weekend)
        }
        
        self.fonts = {
            'header': Font(color='FFFFFF', bold=True, size=12),
            'day_name': Font(bold=True, size=10, color='333333'),
            'day_number': Font(bold=True, size=11, color='2F4F4F'),  # Unified dark gray for all date numbers
            'assignment': Font(size=9, bold=True),
            'blocked': Font(color='000000', bold=True, size=9),  # Black font for blocked cells
        }
        
        self.borders = {
            'thin': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def write_schedule_workbook(self, year: int, months: list[int], schedules: dict, summary: list, employees_data: list = None):
        """
        Write schedule data to proper Excel workbook
        Creates real Excel file with multiple sheets and formatting
        """
        # Create filename
        month_range = f"{min(months):02d}-{max(months):02d}" if len(months) > 1 else f"{months[0]:02d}"
        filename = f"Schedule_{year}_{month_range}.xlsx"
        filepath = self.output_dir / filename
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create summary sheet first
        self._create_summary_sheet(wb, summary)
        
        # Create monthly sheets
        for month in months:
            month_key = f"{year}-{month:02d}"
            month_schedule = schedules.get(month_key, {})
            self._create_month_sheet(
                wb, year, month,
                month_schedule,
                employees_data or []
            )
        
        # Save workbook
        wb.save(filepath)
        print(f"✓ Excel workbook created: {filepath}")
        return filepath
    
    def _create_month_sheet(self, wb: Workbook, year: int, month: int, schedule: dict, employees_data: list):
        """Create matrix layout: employees as rows, dates as columns"""
        month_name = calendar.month_name[month]
        ws = wb.create_sheet(title=f"{month:02d}_{month_name}")
        
        # Title
        ws['A1'] = f"{month_name} {year}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Get all days in the month
        days_in_month = calendar.monthrange(year, month)[1]
        all_dates = []
        for day in range(1, days_in_month + 1):
            all_dates.append(date(year, month, day))
        
        # Create day name headers (row 2)
        ws['A2'] = "Employee Name"
        ws['A2'].fill = self.colors['header']
        ws['A2'].font = self.fonts['header']
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:A3')  # Merge employee name cell across two rows
        
        day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        day_colors = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
        
        for col, date_obj in enumerate(all_dates, 2):  # Start from column B (2)
            # Day name in row 2
            day_name = day_names[date_obj.weekday()]
            day_cell = ws.cell(row=2, column=col, value=day_name)
            day_cell.fill = self.colors[day_colors[date_obj.weekday()]]
            day_cell.font = self.fonts['day_name']
            day_cell.alignment = Alignment(horizontal='center')
            
            # Date number in row 3
            date_cell = ws.cell(row=3, column=col, value=date_obj.day)
            date_cell.fill = self.colors[day_colors[date_obj.weekday()]]
            date_cell.font = self.fonts['day_number']  # Unified color for all date numbers
            date_cell.alignment = Alignment(horizontal='center')
        
        # Add total columns (merge across rows 2-3)
        total_start_col = len(all_dates) + 2
        total_headers = ['WD', 'Th', 'WE', 'B', 'Total']
        for i, header in enumerate(total_headers):
            cell = ws.cell(row=2, column=total_start_col + i, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
            cell.alignment = Alignment(horizontal='center')
            ws.merge_cells(f'{get_column_letter(total_start_col + i)}2:{get_column_letter(total_start_col + i)}3')
        
        # Get blocked days for all employees
        blocked_days = self._get_blocked_days(year, month, employees_data)
        
        # Create employee rows (starting from row 4 now)
        row = 4
        for emp_data in employees_data:
            emp_name = emp_data.get('name', '')
            
            # Employee name column
            name_cell = ws.cell(row=row, column=1, value=emp_name)
            name_cell.font = Font(bold=True)
            name_cell.border = self.borders['thin']
            
            # Count assignments for totals
            emp_counts = {'WD': 0, 'Th': 0, 'WE': 0, 'B': 0}
            
            # Fill date columns
            for col, date_obj in enumerate(all_dates, 2):
                date_str = date_obj.isoformat()
                cell = ws.cell(row=row, column=col)
                cell.border = self.borders['thin']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Check if employee is blocked this day
                is_blocked = date_str in blocked_days and emp_name in blocked_days[date_str]
                
                # Check if employee has assignment this day
                assignment = None
                for day_date, day_assignments in schedule.items():
                    if day_date == date_str:
                        for duty, assigned_emp in day_assignments.items():
                            if assigned_emp == emp_name:
                                assignment = duty
                                emp_counts[duty] += 1
                                break
                        break
                
                # Set cell content and color (moved inside the loop!)
                if is_blocked:
                    cell.value = "-"
                    cell.fill = self.colors['blocked']
                    cell.font = self.fonts['blocked']  # Black font for blocked cells
                elif assignment:
                    cell.value = assignment
                    cell.fill = self.colors[assignment]
                    cell.font = self.fonts['assignment']
                else:
                    # Empty day - light gray background
                    cell.fill = self.colors['empty']
            
            # Add total columns
            total_points = (emp_counts['WD'] * 1.0 + emp_counts['Th'] * 1.5 + 
                          emp_counts['WE'] * 2.0 + emp_counts['B'] * 0.5)
            
            totals = [emp_counts['WD'], emp_counts['Th'], emp_counts['WE'], emp_counts['B'], total_points]
            for i, total_val in enumerate(totals):
                cell = ws.cell(row=row, column=total_start_col + i, value=total_val)
                cell.border = self.borders['thin']
                cell.alignment = Alignment(horizontal='center')
                if i == len(totals) - 1:  # Total points column
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Add totals row
        total_row = row
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        
        # Calculate column totals
        for col in range(2, len(all_dates) + 2):
            # Count assignments per day
            date_obj = all_dates[col - 2]
            date_str = date_obj.isoformat()
            day_assignments = schedule.get(date_str, {})
            assignment_count = len(day_assignments)
            
            cell = ws.cell(row=total_row, column=col, value=assignment_count)
            cell.font = Font(bold=True)
            cell.border = self.borders['thin']
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 20  # Employee names
        for col in range(2, len(all_dates) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 4  # Date columns
        for i in range(len(total_headers)):
            ws.column_dimensions[get_column_letter(total_start_col + i)].width = 8  # Total columns
    
    def _create_summary_sheet(self, wb: Workbook, summary: list):
        """Create the summary sheet with formatting"""
        ws = wb.create_sheet(title="Summary", index=0)  # Insert at beginning
        
        if not summary:
            ws['A1'] = "No summary data available"
            return
        
        # Headers
        headers = list(summary[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, row_data in enumerate(summary, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                cell.border = self.borders['thin']
                
                # Highlight TOTAL row
                if row_data.get('Employee') == 'TOTAL':
                    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _get_blocked_days(self, year: int, month: int, employees_data: list) -> dict:
        """Get blocked days for all employees for the given month"""
        blocked_days = {}
        
        for emp_data in employees_data:
            emp_name = emp_data.get('name', '')
            
            # Get blocked days for this employee
            emp_blocked = set()
            
            # Single blocked days
            for blocked_day in emp_data.get('blocked_days', []):
                if blocked_day.startswith(f"{year}-{month:02d}"):
                    emp_blocked.add(blocked_day)
            
            # Blocked ranges
            for range_data in emp_data.get('blocked_ranges', []):
                try:
                    start_date = date.fromisoformat(range_data['start'])
                    end_date = date.fromisoformat(range_data['end'])
                    
                    # Check if range overlaps with this month
                    month_start = date(year, month, 1)
                    month_end = date(year, month, calendar.monthrange(year, month)[1])
                    
                    if start_date <= month_end and end_date >= month_start:
                        # Add all days in the range that fall in this month
                        current = max(start_date, month_start)
                        end = min(end_date, month_end)
                        
                        while current <= end:
                            emp_blocked.add(current.isoformat())
                            current = current + timedelta(days=1)
                            
                except (ValueError, KeyError):
                    continue
            
            # Sabbath blocking
            if emp_data.get('observes_sabbath'):
                for day in range(1, calendar.monthrange(year, month)[1] + 1):
                    day_date = date(year, month, day)
                    if day_date.weekday() in (4, 5):  # Friday, Saturday
                        emp_blocked.add(day_date.isoformat())
            
            # Add to blocked_days dict
            for blocked_day in emp_blocked:
                if blocked_day not in blocked_days:
                    blocked_days[blocked_day] = []
                blocked_days[blocked_day].append(emp_name)
        
        return blocked_days


def create_excel_output(year: int, months: list[int], output_dir: str = None):
    """
    Create Excel output for a period by reading existing JSON schedules and summary
    """
    from src.config import SUMMARIES_DIR, EMPLOYEES_FILE
    
    if output_dir is None:
        month_range = f"{min(months):02d}-{max(months):02d}" if len(months) > 1 else f"{months[0]:02d}"
        output_dir = Path(SUMMARIES_DIR) / f"{year}_{month_range}"
    
    output_dir = Path(output_dir)
    
    # Load employee data for blocked days visualization
    employees_data = []
    try:
        with open(EMPLOYEES_FILE, 'r', encoding='utf-8') as f:
            employees_data = json.load(f)
    except FileNotFoundError:
        print("Warning: employees.json not found, blocked days won't be shown")
    
    # Load schedules
    schedules = {}
    for month in months:
        schedule_file = output_dir / f"schedule_{year}-{month:02d}.json"
        if schedule_file.exists():
            with open(schedule_file, 'r', encoding='utf-8') as f:
                schedules[f"{year}-{month:02d}"] = json.load(f)
    
    # Load summary
    summary = []
    month_range = f"{min(months):02d}-{max(months):02d}" if len(months) > 1 else f"{months[0]:02d}"
    summary_file = output_dir / f"summary_{year}_{month_range}.csv"
    if summary_file.exists():
        lines = summary_file.read_text(encoding='utf-8').strip().split('\n')
        if len(lines) > 1:
            headers = lines[0].split(',')
            for line in lines[1:]:
                values = line.split(',')
                row = dict(zip(headers, values))
                summary.append(row)
    
    # Create Excel output
    writer = ExcelWriter(str(output_dir))
    return writer.write_schedule_workbook(year, months, schedules, summary, employees_data)
